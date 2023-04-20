from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from datetime import datetime
from pathlib import Path
import os

inventoryExcel = input("재고 파일을 입력해주세요: ")

excelFile = Path(inventoryExcel).exists()

if not excelFile:
    print("파일이 존재하지 않습니다.")
    os.system("pause")

try:
    inventory_wb = load_workbook(inventoryExcel, data_only=True)
    inventory_ws = inventory_wb['Sheet']
except:
    print("파일을 열 수 없습니다.")

todayFormat1 = datetime.today().strftime("%Y%m%d")
todayFormat2 = datetime.today().strftime("%Y-%m-%d")
dirName = todayFormat1 + "_발주서"

# 폴더 생성
if not Path(dirName).exists():
    os.mkdir(dirName)

def nvl(v):
    return "" if v is None else v

def purchaseOrder(row):
    pdCount = len(row)

    try:
        order_wb = load_workbook("purchase_order.xlsx", data_only=True)
        order_ws = order_wb["휴네스"]
    except:
        print("파일을 열 수 없습니다.")

    # 제품 수 만큼 행 추가
    order_ws.insert_rows(13, pdCount)
    supplierName = None

    for x in range(pdCount):
        order_ws["B3"] = todayFormat1 + "_" + str(orderCount)
        order_ws["B4"] = todayFormat2
        
        if supplierName == None:
            supplierName = row[x]["supplier"]

        order_ws["A9"] = supplierName + " 귀하"

        # order_ws["A"+str(13+x)] = row[x]["supplier"]
        order_ws["B"+str(13+x)] = row[x]["product"]
        order_ws["C"+str(13+x)] = row[x]["stockUnit"]
        order_ws["D"+str(13+x)] = int(row[x]["lowInventory"])
        
        # row 높이 조정
        order_ws.row_dimensions[13 + x].height = 25

        # 셀 스타일 적용
        for y in range(ord("a"), ord("g")+1):
            order_ws[chr(y) + str(13 + x)].border = Border(top = Side(border_style='thin', color='000000'),
                                                    right = Side(border_style='thin', color='000000'),
                                                    bottom = Side(border_style='thin', color='000000'),
                                                    left = Side(border_style='thin', color='000000'))

    order_ws.merge_cells("A13:A" + str(13 + pdCount-1))
    order_wb.save(dirName + "/" + todayFormat1 + "_" + supplierName + ".xlsx")
    order_wb.close()
    
supplier = None
all_values = []
orderCount = 0

for idx, row in enumerate(inventory_ws.rows):
    # 첫째줄은 통과
    if idx == 0 or row[6].value == None or 0 > row[1].value.find(")"):
        continue     

    # 이전값과 비교하기 위해
    temp = nvl(row[1].value[0: row[1].value.find(")")]).strip()
    
    # 업체별로 문서를 생성하기 위한 조건
    if (supplier != None) and (supplier != temp):
        orderCount += 1
        # print(all_values)
        try:
            purchaseOrder(all_values)
        except PermissionError:
            print("===> 업체명: " + supplier + " 파일이 열려 있어 생성할 수 없습니다. ")

        all_values = []
                
        print(str(orderCount) + "." + supplier)
        # break

    supplier = temp

    data = {}
    data["supplier"] = temp
    data["product"] = nvl(row[1].value[row[1].value.find(")") + 1: len(row[1].value)]).strip()
    data["stockUnit"] = nvl(row[3].value).strip()
    data["lowInventory"] = row[6].value

    all_values.append(data)
 
    # print("업체: " + data["supplier"] + ", 재고단위: " + data["stockUnit"] + ", 제품: " + data["product"] + ", 부족재고: " + data["lowInventory"])
else:
    print("발주서 생성이 완료되었습니다.")
    os.system("pause")
